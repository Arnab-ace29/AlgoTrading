"""
Build the master stock data inventory Excel.

Columns per stock:
  - Symbol, Index Membership, Cap Category
  - Market Cap (INR Cr), Avg Daily Turnover (INR Cr), Sector, Industry
  - 5-min: Available, Start Date, End Date, Bar Count, Trading Days, Fill %
  - 1-day: Available, Start Date, End Date, Bar Count, Trading Days
  - F&O Eligible, Strategy Filter Pass (mkt cap ≥ 1000cr + turnover ≥ 50cr/day)
  - Upstox Instrument Key

Run:
    python scripts/build_data_excel.py
    python scripts/build_data_excel.py --no-yfinance   # skip market cap fetch (fast)
    python scripts/build_data_excel.py --output my.xlsx

Output: data/stock_master.xlsx  (updated every run)
"""

from __future__ import annotations

import argparse
import json
import sqlite3
import sys
import time
from pathlib import Path
from datetime import datetime

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
from loguru import logger

DB_PATH         = ROOT / "data" / "algo_trading.sqlite"
UNIVERSES_FILE  = ROOT / "config" / "universes.json"
DEFAULT_OUTPUT  = ROOT / "data" / "stock_master.xlsx"
YF_CACHE_FILE   = ROOT / "data" / "yfinance_info_cache.json"
YF_CACHE_TTL    = 7 * 86400   # re-use cached market caps for a week

# ── Helpers ────────────────────────────────────────────────────────────────────

def load_universes() -> dict[str, set[str]]:
    with open(UNIVERSES_FILE, encoding="utf-8") as f:
        raw = json.load(f)
    return {k: set(v) for k, v in raw.items()}


def get_db_coverage(conn: sqlite3.Connection) -> dict[str, dict]:
    """Fetch bar count, date range per symbol per timeframe."""
    rows = conn.execute(
        """SELECT symbol, timeframe, COUNT(*) as bars,
                  MIN(timestamp) as first_ts, MAX(timestamp) as last_ts
           FROM minute_candles
           GROUP BY symbol, timeframe"""
    ).fetchall()
    coverage: dict[str, dict] = {}
    for sym, tf, bars, first_ts, last_ts in rows:
        if sym not in coverage:
            coverage[sym] = {}
        coverage[sym][tf] = {
            "bars":  bars,
            "first": str(first_ts)[:10] if first_ts else None,
            "last":  str(last_ts)[:10]  if last_ts  else None,
        }
    return coverage


def get_batch_daily_metrics(conn: sqlite3.Connection, last_n_days: int = 30) -> dict[str, dict]:
    """
    Single-pass batch query: compute avg daily turnover and ATR% for ALL symbols.
    Runs one SQL query instead of one-per-symbol — 100x faster on 23M-row table.
    Returns {symbol: {turnover_cr, atr_pct}}
    """
    logger.info("  Computing daily turnover + ATR% for all symbols (single batch query)...")
    # Get daily OHLCV aggregates for the last 60 trading days per symbol
    rows = conn.execute(
        """SELECT symbol, DATE(timestamp) as d,
                  SUM(volume * close)   as turnover,
                  MAX(high) - MIN(low)  as range_abs,
                  AVG(close)            as avg_close
           FROM minute_candles
           WHERE timeframe='5min'
           GROUP BY symbol, d
           ORDER BY symbol, d DESC"""
    ).fetchall()

    from collections import defaultdict
    import statistics

    sym_data: dict[str, list] = defaultdict(list)
    for sym, d, turnover, range_abs, avg_close in rows:
        sym_data[sym].append((turnover or 0, range_abs, avg_close))

    result = {}
    for sym, day_rows in sym_data.items():
        recent = day_rows[:last_n_days]  # already sorted desc
        if not recent:
            result[sym] = {"turnover_cr": 0.0, "atr_pct": 0.0}
            continue
        avg_turnover = sum(r[0] for r in recent) / len(recent) / 1e7
        atrs = [r[1] / r[2] * 100 for r in recent if r[2] and r[2] > 0]
        avg_atr = sum(atrs) / len(atrs) if atrs else 0.0
        result[sym] = {
            "turnover_cr": round(avg_turnover, 2),
            "atr_pct":     round(avg_atr, 2),
        }
    logger.info(f"  Batch metrics computed for {len(result)} symbols.")
    return result


def _load_yf_cache() -> dict[str, dict]:
    """Load the yfinance info cache if it exists and is fresh (< TTL)."""
    if not YF_CACHE_FILE.exists():
        return {}
    if time.time() - YF_CACHE_FILE.stat().st_mtime > YF_CACHE_TTL:
        logger.info("  yfinance cache is stale (> 7d); will refresh.")
        return {}
    try:
        return json.loads(YF_CACHE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def fetch_yfinance_info(symbols: list[str], refresh: bool = False) -> dict[str, dict]:
    """
    Market cap, sector, industry, name, ISIN per symbol from yfinance.

    Cached to data/yfinance_info_cache.json (7-day TTL): only symbols missing from
    the cache are fetched, so adding a few names to the universe costs a few calls,
    not a 15-minute full re-fetch. Pass refresh=True to force a full re-fetch.
    """
    try:
        import yfinance as yf
    except ImportError:
        logger.warning("yfinance not installed; skip market cap fetch")
        return {}

    cache = {} if refresh else _load_yf_cache()
    to_fetch = [s for s in symbols if s not in cache]
    logger.info(f"  yfinance: {len(cache)} cached, {len(to_fetch)} to fetch")

    info: dict[str, dict] = {s: cache[s] for s in symbols if s in cache}
    for i, sym in enumerate(to_fetch, 1):
        ticker_sym = f"{sym}.NS"
        try:
            d = yf.Ticker(ticker_sym).info
            mkt_cap_cr = round((d.get("marketCap") or 0) / 1e7, 0)
            info[sym] = {
                "market_cap_cr": mkt_cap_cr,
                "sector":        d.get("sector", ""),
                "industry":      d.get("industry", ""),
                "name":          d.get("longName", d.get("shortName", sym)),
                "isin":          d.get("isin", ""),
            }
            if i % 50 == 0:
                logger.info(f"  yfinance: {i}/{len(to_fetch)} fetched")
            time.sleep(0.3)
        except Exception as e:
            logger.debug(f"  yfinance error for {ticker_sym}: {e}")
            info[sym] = {"market_cap_cr": 0, "sector": "", "industry": "", "name": sym, "isin": ""}

    # Persist the merged cache (existing entries + newly fetched).
    try:
        merged = {**cache, **info}
        YF_CACHE_FILE.write_text(json.dumps(merged), encoding="utf-8")
    except Exception as e:
        logger.warning(f"Could not write yfinance cache: {e}")
    return info


def get_cap_category(sym: str, universes: dict[str, set[str]], mkt_cap_cr: float) -> str:
    if sym in universes.get("nifty50", set()):
        return "Large Cap (Nifty 50)"
    if sym in universes.get("nifty100", set()):
        return "Large Cap (Nifty 100)"
    if mkt_cap_cr >= 20000:
        return "Large Cap"
    if mkt_cap_cr >= 5000:
        return "Mid Cap"
    if mkt_cap_cr >= 1000:
        return "Small Cap"
    if mkt_cap_cr > 0:
        return "Micro/Nano Cap"
    # Fallback by position in nifty_total (top 250 → mid, rest → small)
    total_list = list(universes.get("nifty_total", []))
    try:
        rank = total_list.index(sym)
        if rank < 100:
            return "Large Cap"
        if rank < 250:
            return "Mid Cap"
        if rank < 500:
            return "Small Cap"
        return "Micro Cap"
    except ValueError:
        return "Unknown"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--no-yfinance", action="store_true", help="Skip yfinance market cap fetch (fast mode)")
    parser.add_argument("--refresh-yf", action="store_true", help="Force full yfinance re-fetch (ignore cache)")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()

    logger.info("Loading universe and DB coverage...")
    universes = load_universes()

    # Build master symbol list: all unique symbols across all universe lists
    all_syms: set[str] = set()
    for lst in universes.values():
        all_syms.update(lst)
    # Also add the index/VIX symbols we store
    extra = {"INDIAVIX", "NIFTY50", "NIFTYBANK", "NIFTYIT", "NIFTYFMCG",
             "NIFTYPHARMA", "NIFTYAUTO", "NIFTYMETAL", "NIFTYREALTY"}
    # But don't include these in the equity rows (filter later)
    all_equity_syms = sorted(all_syms)
    logger.info(f"  Total symbols in universe: {len(all_equity_syms)}")

    conn = sqlite3.connect(str(DB_PATH), timeout=30)
    coverage = get_db_coverage(conn)
    logger.info(f"  Symbols with any data in DB: {len(coverage)}")

    # yfinance enrichment (cached — only new symbols are fetched)
    yf_info: dict[str, dict] = {}
    if not args.no_yfinance:
        logger.info(f"yfinance info for {len(all_equity_syms)} symbols (cache-aware)...")
        yf_info = fetch_yfinance_info(all_equity_syms, refresh=args.refresh_yf)
    else:
        logger.info("Skipping yfinance (--no-yfinance)")

    # Latest trading day present in the DB — reference for the staleness flag.
    row = conn.execute("SELECT MAX(timestamp) FROM minute_candles WHERE timeframe='1day'").fetchone()
    db_last_date = str(row[0])[:10] if row and row[0] else None

    # Single batch query for turnover + ATR (avoids 748 per-symbol queries)
    daily_metrics = get_batch_daily_metrics(conn)

    logger.info("Computing DB metrics per symbol...")
    rows = []
    for sym in all_equity_syms:
        cov = coverage.get(sym, {})
        c5  = cov.get("5min",  {"bars": 0, "first": None, "last": None})
        c1d = cov.get("1day",  {"bars": 0, "first": None, "last": None})

        # Trading days estimate
        trading_days_5m = c5["bars"] // 75 if c5["bars"] else 0  # 75 bars/day for 5min
        fill_pct_5m = round(trading_days_5m / 500 * 100, 1) if trading_days_5m else 0  # vs 2yr expected

        # Turnover + ATR from batch query
        m = daily_metrics.get(sym, {"turnover_cr": 0.0, "atr_pct": 0.0})
        avg_turnover = m["turnover_cr"]
        atr_pct      = m["atr_pct"]

        yf  = yf_info.get(sym, {})
        mkt_cap = yf.get("market_cap_cr", 0)
        cap_cat = get_cap_category(sym, universes, mkt_cap)

        # Index membership flags
        in_n50   = sym in universes.get("nifty50", set())
        in_n100  = sym in universes.get("nifty100", set())
        in_ntot  = sym in universes.get("nifty_total", set())
        in_fo    = sym in universes.get("fo_eligible", set())

        # Strategy filter
        filter_pass = (mkt_cap >= 1000 or in_n100) and avg_turnover >= 50

        # Two-tier universe (for the strategy): A = Nifty 100 (tight cost, full size),
        # B = liquid mid/large that pass the filter, — = doesn't qualify.
        if in_n100:
            tier = "A · Nifty100"
        elif filter_pass:
            tier = "B · Mid/Liquid"
        else:
            tier = "—"

        # History depth (explains low bar counts: recent IPOs aren't a data problem).
        if c1d["bars"] >= 450:
            history = "Full (2yr)"
        elif c1d["bars"] > 0:
            history = "Partial (IPO/new)"
        else:
            history = "None"

        # Staleness: how far the symbol's last 1-day bar lags the DB's latest day.
        days_stale = ""
        data_current = "N"
        if c1d["last"] and db_last_date:
            try:
                lag = (datetime.fromisoformat(db_last_date) - datetime.fromisoformat(c1d["last"])).days
                days_stale = lag
                data_current = "Y" if lag <= 1 else "N"   # 1-day tolerance (download timing)
            except ValueError:
                pass

        rows.append({
            "Symbol":               sym,
            "Company Name":         yf.get("name", sym),
            "ISIN":                 yf.get("isin", ""),
            "Sector":               yf.get("sector", ""),
            "Industry":             yf.get("industry", ""),
            "Cap Category":         cap_cat,
            "Universe Tier":        tier,
            "Market Cap (₹ Cr)":   mkt_cap,
            "Avg Daily Turnover (₹ Cr)": avg_turnover,
            "ATR %":                atr_pct,
            "Nifty 50":             "Y" if in_n50  else "",
            "Nifty 100":            "Y" if in_n100 else "",
            "Nifty Total Market":   "Y" if in_ntot else "",
            "F&O Eligible":         "Y" if in_fo   else "",
            "Strategy Filter Pass": "Y" if filter_pass else "N",
            "History":              history,
            "Data Current?":        data_current,
            "Days Stale":           days_stale,
            "5min: Available":      "Y" if c5["bars"] > 0 else "N",
            "5min: Start Date":     c5["first"] or "",
            "5min: End Date":       c5["last"]  or "",
            "5min: Bar Count":      c5["bars"],
            "5min: Trading Days":   trading_days_5m,
            "5min: Fill % (2yr)":   fill_pct_5m,
            "1day: Available":      "Y" if c1d["bars"] > 0 else "N",
            "1day: Start Date":     c1d["first"] or "",
            "1day: End Date":       c1d["last"]  or "",
            "1day: Bar Count":      c1d["bars"],
            "1day: Trading Days":   c1d["bars"],
            "OHLCV Source":         "Upstox",
            "Fundamentals Source":  "yfinance" if (yf.get("market_cap_cr") or 0) > 0 else "—",
            "Last Updated":         datetime.now().strftime("%Y-%m-%d %H:%M"),
        })

    conn.close()
    df = pd.DataFrame(rows)

    # Sort: Nifty 50 first, then Nifty 100, then rest (by bar count desc)
    df["_sort_key"] = df.apply(lambda r: (
        0 if r["Nifty 50"] == "Y" else (1 if r["Nifty 100"] == "Y" else 2)
    ), axis=1)
    df = df.sort_values(["_sort_key", "5min: Bar Count"], ascending=[True, False]).drop(columns=["_sort_key"])

    # ── Write Excel ───────────────────────────────────────────────────────────
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"Writing Excel → {output_path}")
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Stock Master")
        ws = writer.sheets["Stock Master"]

        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        # Column widths by NAME (robust to column additions/reordering).
        name_widths = {
            "Symbol": 13, "Company Name": 28, "ISIN": 14, "Sector": 18, "Industry": 22,
            "Cap Category": 20, "Universe Tier": 16, "Market Cap (₹ Cr)": 15,
            "Avg Daily Turnover (₹ Cr)": 22, "ATR %": 8,
            "Nifty 50": 9, "Nifty 100": 9, "Nifty Total Market": 16, "F&O Eligible": 12,
            "Strategy Filter Pass": 18, "History": 17, "Data Current?": 13, "Days Stale": 11,
        }
        default_w = 14
        for idx, col in enumerate(df.columns, start=1):
            letter = get_column_letter(idx)
            ws.column_dimensions[letter].width = name_widths.get(col, default_w)

        # Header row: bold + light blue fill
        header_fill = PatternFill("solid", fgColor="BDD7EE")
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.fill  = header_fill
            cell.font  = bold
            cell.alignment = Alignment(wrap_text=True, horizontal="center")

        # Color-code Strategy Filter Pass column
        green_fill  = PatternFill("solid", fgColor="C6EFCE")
        red_fill    = PatternFill("solid", fgColor="FFC7CE")
        yellow_fill = PatternFill("solid", fgColor="FFEB9C")

        filter_col_idx  = df.columns.get_loc("Strategy Filter Pass") + 1
        avail_5m_idx    = df.columns.get_loc("5min: Available") + 1
        current_col_idx = df.columns.get_loc("Data Current?") + 1

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            filt_cell    = row[filter_col_idx - 1]
            avail_cell   = row[avail_5m_idx - 1]
            current_cell = row[current_col_idx - 1]
            if filt_cell.value == "Y":
                filt_cell.fill = green_fill
            elif filt_cell.value == "N":
                filt_cell.fill = red_fill
            if avail_cell.value == "N":
                avail_cell.fill = yellow_fill
            # Stale data → red flag so it's obvious what needs a re-download.
            if current_cell.value == "N":
                current_cell.fill = red_fill
            elif current_cell.value == "Y":
                current_cell.fill = green_fill

        # Freeze header row
        ws.freeze_panes = "A2"

        # Summary sheet
        n_total   = len(df)
        n_5m      = (df["5min: Available"] == "Y").sum()
        n_1d      = (df["1day: Available"] == "Y").sum()
        n_filter  = (df["Strategy Filter Pass"] == "Y").sum()
        n_n100    = (df["Nifty 100"] == "Y").sum()
        n_tierA   = (df["Universe Tier"] == "A · Nifty100").sum()
        n_tierB   = (df["Universe Tier"] == "B · Mid/Liquid").sum()
        n_fo      = (df["F&O Eligible"] == "Y").sum()
        n_full    = (df["History"] == "Full (2yr)").sum()
        n_partial = (df["History"] == "Partial (IPO/new)").sum()
        n_current = (df["Data Current?"] == "Y").sum()
        n_stale   = n_total - n_current

        summary_data = [
            ["Metric", "Value"],
            ["Total symbols in universe",    n_total],
            ["— Tier A (Nifty 100)",         n_tierA],
            ["— Tier B (Mid/Liquid, passes filter)", n_tierB],
            ["— F&O eligible",               n_fo],
            ["Pass strategy filter (≥1000cr mktcap + ≥50cr/day vol)", n_filter],
            ["", ""],
            ["Symbols with 5-min data",      n_5m],
            ["Symbols with 1-day data",      n_1d],
            ["— Full history (~2yr)",        n_full],
            ["— Partial history (IPO/new)",  n_partial],
            ["", ""],
            ["DB latest trading day",        db_last_date or "—"],
            ["Data current (up to latest day)", n_current],
            ["Data STALE (needs re-download)",  n_stale],
            ["", ""],
            ["Generated at",                  datetime.now().strftime("%Y-%m-%d %H:%M")],
        ]
        ws_sum = writer.book.create_sheet("Summary")
        for r_data in summary_data:
            ws_sum.append(r_data)
        for cell in ws_sum[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        ws_sum.column_dimensions["A"].width = 52
        ws_sum.column_dimensions["B"].width = 18

        # ── Indices & Macro sheet ─────────────────────────────────────────────
        INDEX_META = {
            # symbol: (description, type, source)
            "INDIAVIX":    ("India VIX (Volatility Index)",          "India VIX",      "yfinance"),
            "NIFTY50_YF":  ("Nifty 50 — yfinance feed",              "India Index",    "yfinance"),
            "SP500":       ("S&P 500 (US large-cap index)",           "Global Index",   "yfinance"),
            "NASDAQ":      ("NASDAQ Composite",                       "Global Index",   "yfinance"),
            "NIFTY50":     ("Nifty 50",                               "India Index",    "Upstox"),
            "NIFTYNEXT50": ("Nifty Next 50",                          "India Index",    "Upstox"),
            "NIFTYBANK":   ("Nifty Bank",                             "Sector Index",   "Upstox"),
            "NIFTYIT":     ("Nifty IT",                               "Sector Index",   "Upstox"),
            "NIFTYFMCG":   ("Nifty FMCG",                            "Sector Index",   "Upstox"),
            "NIFTYPHARMA": ("Nifty Pharma",                           "Sector Index",   "Upstox"),
            "NIFTYAUTO":   ("Nifty Auto",                             "Sector Index",   "Upstox"),
            "NIFTYMETAL":  ("Nifty Metal",                            "Sector Index",   "Upstox"),
            "NIFTYREALTY": ("Nifty Realty",                           "Sector Index",   "Upstox"),
            "NIFTYINFRA":  ("Nifty Infrastructure (key invalid)",     "Sector Index",   "Upstox"),
        }

        conn2 = sqlite3.connect(str(DB_PATH), timeout=30)
        idx_rows = []
        for sym, (desc, itype, src) in INDEX_META.items():
            r = conn2.execute(
                "SELECT COUNT(*), MIN(timestamp), MAX(timestamp) FROM minute_candles WHERE symbol=?",
                (sym,)
            ).fetchone()
            bars, first_ts, last_ts = r
            idx_rows.append({
                "Symbol":      sym,
                "Description": desc,
                "Type":        itype,
                "Source":      src,
                "Timeframe":   "1day",
                "Start Date":  str(first_ts)[:10] if first_ts else "—",
                "End Date":    str(last_ts)[:10]  if last_ts  else "—",
                "Bar Count":   bars or 0,
                "Available":   "Y" if bars else "N",
            })
        conn2.close()

        df_idx = pd.DataFrame(idx_rows)
        df_idx.to_excel(writer, index=False, sheet_name="Indices & Macro")
        ws_idx = writer.sheets["Indices & Macro"]

        # Header formatting
        for cell in ws_idx[1]:
            cell.fill  = PatternFill("solid", fgColor="D9E1F2")
            cell.font  = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws_idx.freeze_panes = "A2"

        col_w_idx = {"A": 16, "B": 42, "C": 16, "D": 12, "E": 12, "F": 14, "G": 14, "H": 12, "I": 12}
        for col_letter, width in col_w_idx.items():
            ws_idx.column_dimensions[col_letter].width = width

        # Colour Available column
        avail_idx_col = df_idx.columns.get_loc("Available") + 1
        for row in ws_idx.iter_rows(min_row=2, max_row=ws_idx.max_row):
            cell = row[avail_idx_col - 1]
            if cell.value == "Y":
                cell.fill = green_fill
            elif cell.value == "N":
                cell.fill = red_fill

    logger.success(f"Excel written: {output_path}  ({n_total} symbols, {n_5m} with 5min data)")
    logger.info(f"  Strategy filter candidates: {n_filter}")


if __name__ == "__main__":
    main()
